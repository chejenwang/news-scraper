"""
TASS 新聞爬蟲 - 改進版（雙翻譯引擎）
支援 googletrans 和 deep-translator 雙引擎
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import time
import random
from typing import List, Dict
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class TASSNewsScraper:
    def __init__(self, headless=True):
        """初始化爬蟲，使用 Selenium"""
        print("正在啟動瀏覽器...")
        
        chrome_options = Options()
        
        if headless:
            chrome_options.add_argument('--headless')  # 無頭模式
        
        # 隱藏警告訊息
        chrome_options.add_argument('--log-level=3')
        chrome_options.add_argument('--disable-logging')
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        # 模擬真實瀏覽器
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        chrome_options.add_argument('--lang=ru-RU')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # 類別名稱對照
        self.categories = {
            'politika': '政治',
            'ekonomika': '經濟',
            'armiya-i-opk': '軍事與國防',
            'mezhdunarodnaya-panorama': '國際全景'
        }
        
        print("✓ 瀏覽器啟動完成\n")
    
    def __del__(self):
        """關閉瀏覽器"""
        try:
            self.driver.quit()
            print("\n✓ 瀏覽器已關閉")
        except:
            pass
    
    def human_delay(self, min_seconds=3, max_seconds=7):
        """模擬人類閱讀延遲"""
        delay = random.uniform(min_seconds, max_seconds)
        print(f"  等待 {delay:.1f} 秒...")
        time.sleep(delay)
    
    def scroll_page(self):
        """模擬人類滾動頁面"""
        total_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_pause = random.uniform(0.5, 1.5)
        
        # 分段滾動
        for i in range(0, total_height, 300):
            self.driver.execute_script(f"window.scrollTo(0, {i});")
            time.sleep(scroll_pause)
    
    def translate_to_chinese(self, text: str, max_retries=3) -> str:
        """翻譯俄文到繁體中文 - 使用多個備用翻譯引擎"""
        if not text or len(text.strip()) == 0:
            return ""
        
        # 方法1: 優先使用 googletrans (免費且較穩定)
        for attempt in range(max_retries):
            try:
                from googletrans import Translator
                translator = Translator()
                
                max_length = 4500
                if len(text) > max_length:
                    chunks = [text[i:i+max_length] for i in range(0, len(text), max_length)]
                    translated_chunks = []
                    for chunk in chunks:
                        result = translator.translate(chunk, src='ru', dest='zh-tw')
                        translated_chunks.append(result.text)
                        time.sleep(random.uniform(1.5, 2.5))
                    return ''.join(translated_chunks)
                else:
                    result = translator.translate(text, src='ru', dest='zh-tw')
                    return result.text
                    
            except Exception as e:
                if attempt < max_retries - 1:
                    wait_time = random.uniform(4, 7)
                    print(f"  googletrans 失敗，等待 {wait_time:.1f} 秒後重試 ({attempt + 1}/{max_retries})...")
                    time.sleep(wait_time)
                else:
                    print(f"  googletrans 失敗: {str(e)[:50]}")
        
        # 方法2: 備用方案 - 使用 deep-translator
        try:
            from deep_translator import GoogleTranslator
            print("  切換到 deep-translator...")
            
            max_length = 4500
            if len(text) > max_length:
                chunks = [text[i:i+max_length] for i in range(0, len(text), max_length)]
                translated_chunks = []
                for chunk in chunks:
                    translator = GoogleTranslator(source='ru', target='zh-TW')
                    result = translator.translate(chunk)
                    translated_chunks.append(result)
                    time.sleep(random.uniform(2, 3))
                return ''.join(translated_chunks)
            else:
                translator = GoogleTranslator(source='ru', target='zh-TW')
                return translator.translate(text)
        except Exception as e:
            print(f"  deep-translator 也失敗: {str(e)[:50]}")
        
        return "[翻譯失敗]"
    
    def get_article_links(self, category_url: str, max_articles: int = 10) -> List[str]:
        """獲取類別頁面中的文章連結"""
        print(f"\n正在獲取文章列表: {category_url}")
        
        try:
            self.driver.get(category_url)
            self.human_delay(3, 5)
            
            # 滾動頁面載入更多內容
            print("  滾動頁面載入內容...")
            self.scroll_page()
            time.sleep(2)
            
            # 獲取頁面源碼
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            article_links = []
            news_links = soup.find_all('a', href=True)
            
            for a_tag in news_links:
                href = a_tag['href']
                
                # TASS 文章 URL 模式
                if re.search(r'/(politika|ekonomika|armiya-i-opk|mezhdunarodnaya-panorama)/\d+', href):
                    # 轉換為完整 URL
                    if href.startswith('http'):
                        full_url = href
                    elif href.startswith('/'):
                        full_url = f"https://tass.ru{href}"
                    else:
                        continue
                    
                    # 去重並確保是 tass.ru 域名
                    if 'tass.ru' in full_url and full_url not in article_links:
                        article_links.append(full_url)
                        print(f"  找到: {full_url}")
                        
                        if len(article_links) >= max_articles:
                            break
            
            # 備用方案：如果沒找到，嘗試更寬鬆的匹配
            if len(article_links) == 0:
                print("  使用備用方法搜尋連結...")
                for a_tag in news_links:
                    href = a_tag.get('href', '')
                    if re.search(r'/\d{7,}', href) and \
                       not any(x in href for x in ['page', 'tag', 'search', 'about']):
                        
                        if href.startswith('/'):
                            full_url = f"https://tass.ru{href}"
                        elif href.startswith('http'):
                            full_url = href
                        else:
                            continue
                        
                        if 'tass.ru' in full_url and full_url not in article_links:
                            article_links.append(full_url)
                            print(f"  找到: {full_url}")
                            
                            if len(article_links) >= max_articles:
                                break
            
            print(f"總共找到 {len(article_links)} 篇文章\n")
            return article_links
            
        except Exception as e:
            print(f"❌ 獲取文章列表錯誤: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def clean_text(self, text: str) -> str:
        """清理文本"""
        if not text:
            return ""
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        text = text.replace('\xa0', ' ')
        return text
    
    def scrape_article(self, url: str) -> Dict:
        """爬取單篇文章"""
        print(f"\n{'='*70}")
        print(f"正在爬取: {url}")
        
        try:
            self.driver.get(url)
            self.human_delay(4, 8)  # 每篇文章等待 4-8 秒
            
            # 滾動頁面確保所有內容載入
            self.scroll_page()
            time.sleep(1)
            
            # 獲取頁面源碼
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            article_data = {
                'url': url,
                'download_datetime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # === 標題 ===
            title = ''
            try:
                title_elem = self.driver.find_element(By.TAG_NAME, 'h1')
                title = self.clean_text(title_elem.text)
            except:
                title_tag = soup.find('h1')
                if title_tag:
                    title = self.clean_text(title_tag.get_text())
            
            if not title:
                meta_title = soup.find('meta', {'property': 'og:title'})
                if meta_title:
                    title = meta_title.get('content', '')
            
            article_data['title_ru'] = title
            print(f"✓ 標題: {title[:60]}...")
            
            # === 日期 ===
            date = ''
            date_selectors = [
                (By.TAG_NAME, 'time'),
                (By.CLASS_NAME, 'date'),
                (By.CLASS_NAME, 'Datetime'),
            ]
            
            for by, value in date_selectors:
                try:
                    date_elem = self.driver.find_element(by, value)
                    date = self.clean_text(date_elem.text)
                    if not date:
                        date = date_elem.get_attribute('datetime') or ''
                    if date:
                        break
                except:
                    continue
            
            # 備用：從 soup 找
            if not date:
                time_tag = soup.find('time')
                if time_tag:
                    date = time_tag.get('datetime', '') or self.clean_text(time_tag.get_text())
            
            article_data['date'] = date
            print(f"✓ 日期: {date if date else '未找到'}")
            
            # === 作者 ===
            author = ''
            try:
                author_elem = self.driver.find_element(By.CLASS_NAME, 'Author')
                author = self.clean_text(author_elem.text)
            except:
                pass
            
            if not author:
                author_tag = soup.find('span', class_=re.compile(r'author', re.I))
                if author_tag:
                    author = self.clean_text(author_tag.get_text())
            
            if not author:
                author = 'TASS'
            
            article_data['author'] = author
            print(f"✓ 作者: {author}")
            
            # === 內文 ===
            content_parts = []
            
            # 嘗試找到文章主體
            try:
                article_body = self.driver.find_element(By.CLASS_NAME, 'text-block')
                paragraphs = article_body.find_elements(By.TAG_NAME, 'p')
                
                for para in paragraphs:
                    text = self.clean_text(para.text)
                    if len(text) > 20:
                        content_parts.append(text)
            except:
                # 備用方案：從 soup 提取
                print("  使用備用方法提取內文...")
                
                article_containers = soup.find_all(['article', 'div'], class_=re.compile(r'text-block|article-text|content|body', re.I))
                
                for container in article_containers:
                    for para in container.find_all(['p', 'h2', 'h3']):
                        if para.find_parent(['aside', 'footer', 'nav', 'header']):
                            continue
                        
                        text = self.clean_text(para.get_text())
                        if len(text) > 20 and text not in content_parts:
                            content_parts.append(text)
                
                # 如果還是找不到，直接找所有 p 標籤
                if len(content_parts) < 2:
                    for p in soup.find_all('p'):
                        if p.find_parent(['nav', 'footer', 'header', 'aside']):
                            continue
                        text = self.clean_text(p.get_text())
                        if len(text) > 30 and text not in content_parts:
                            content_parts.append(text)
            
            content = '\n\n'.join(content_parts)
            article_data['content_ru'] = content
            
            print(f"✓ 內文長度: {len(content)} 字符，{len(content_parts)} 段落")
            
            # 警告
            if not content:
                print("⚠️ 警告: 未能提取到內文")
            if not date:
                print("⚠️ 警告: 未能提取到日期")
            
            return article_data
            
        except Exception as e:
            print(f"❌ 爬取錯誤: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def scrape_category(self, category_url: str, max_articles: int = 10) -> List[Dict]:
        """爬取單一類別的新聞"""
        category_name = category_url.split('/')[-1]
        print(f"\n{'#'*70}")
        print(f"# 開始爬取類別: {self.categories.get(category_name, category_name)}")
        print(f"# URL: {category_url}")
        print(f"{'#'*70}")
        
        # 獲取文章連結
        article_links = self.get_article_links(category_url, max_articles)
        
        if not article_links:
            print("⚠️ 未找到文章連結")
            return []
        
        # 爬取文章
        articles = []
        for i, link in enumerate(article_links, 1):
            print(f"\n進度: {i}/{len(article_links)}")
            
            article = self.scrape_article(link)
            if article:
                article['category'] = self.categories.get(category_name, category_name)
                articles.append(article)
            
            # 每 3 篇休息 10-15 秒
            if i % 3 == 0 and i < len(article_links):
                print("\n休息一下，避免被偵測...")
                rest_time = random.uniform(10, 15)
                print(f"  休息 {rest_time:.1f} 秒")
                time.sleep(rest_time)
        
        return articles
    
    def scrape_all_categories(self, urls: List[str], max_per_category: int = 10) -> List[Dict]:
        """爬取所有類別"""
        all_articles = []
        
        for i, url in enumerate(urls, 1):
            print(f"\n\n{'='*70}")
            print(f"處理類別 {i}/{len(urls)}")
            print(f"{'='*70}")
            
            articles = self.scrape_category(url, max_per_category)
            all_articles.extend(articles)
            
            # 類別之間休息 20-30 秒
            if i < len(urls):
                rest_time = random.uniform(20, 30)
                print(f"\n完成此類別，休息 {rest_time:.1f} 秒後繼續...")
                time.sleep(rest_time)
        
        return all_articles
    
    def translate_articles(self, articles: List[Dict]):
        """批次翻譯文章 - 改進版"""
        print(f"\n{'='*70}")
        print("開始翻譯文章...")
        print(f"提示：翻譯可能需要較長時間，請耐心等待")
        print(f"{'='*70}\n")
        
        successful = 0
        failed = 0
        
        for i, article in enumerate(articles, 1):
            print(f"\n{'='*70}")
            print(f"翻譯進度: {i}/{len(articles)}")
            print(f"{'='*70}")
            print(f"標題: {article['title_ru'][:50]}...")
            
            # 翻譯標題
            if article.get('title_ru'):
                article['title_zh'] = self.translate_to_chinese(article['title_ru'])
                if article['title_zh'] != "[翻譯失敗]":
                    print(f"✓ 標題: {article['title_zh'][:50]}...")
                    successful += 1
                else:
                    print(f"✗ 標題翻譯失敗")
                    failed += 1
                time.sleep(random.uniform(2, 4))  # 增加延遲避免被限流
            else:
                article['title_zh'] = ''
            
            # 翻譯內文
            if article.get('content_ru'):
                print(f"翻譯內文 ({len(article['content_ru'])} 字符)...")
                article['content_zh'] = self.translate_to_chinese(article['content_ru'])
                
                # 檢查是否成功
                if article['content_zh'] == "[翻譯失敗]":
                    print(f"✗ 內文翻譯失敗")
                else:
                    print(f"✓ 內文翻譯完成 ({len(article['content_zh'])} 字符)")
                
                time.sleep(random.uniform(3, 6))  # 更長的延遲
            else:
                article['content_zh'] = ''
            
            # 每5篇休息一下
            if i % 5 == 0 and i < len(articles):
                rest_time = random.uniform(15, 20)
                print(f"\n{'🔄'*20}")
                print(f"已翻譯 {i}/{len(articles)} 篇")
                print(f"休息 {rest_time:.1f} 秒避免限流...")
                print(f"{'🔄'*20}")
                time.sleep(rest_time)
        
        print(f"\n{'='*70}")
        print(f"翻譯統計: 成功 {successful}, 失敗 {failed}")
        print(f"{'='*70}")
    
    def save_to_excel(self, articles: List[Dict], filename: str = 'tass_news.xlsx'):
        """儲存為Excel並設定自動換行"""
        data = []
        for article in articles:
            data.append({
                '日期': article.get('date', ''),
                '作者': article.get('author', ''),
                '標題_俄文': article.get('title_ru', ''),
                '標題_中文': article.get('title_zh', ''),
                '內文_俄文': article.get('content_ru', ''),
                '內文_中文': article.get('content_zh', ''),
                '網址': article.get('url', ''),
                '下載時間': article.get('download_datetime', ''),
                '分類': article.get('category', ''),
            })
        
        df = pd.DataFrame(data)
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # 格式設定
        wb = load_workbook(filename)
        ws = wb.active
        
        column_widths = {
            'A': 18, 'B': 20, 'C': 60, 'D': 60,
            'E': 100, 'F': 100, 'G': 50, 'H': 20, 'I': 15,
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 設定自動換行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='top',
                    horizontal='left'
                )
        
        # 標題行
        for cell in ws[1]:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical='center',
                horizontal='center'
            )
        
        wb.save(filename)
        print(f"\n✓ 已儲存Excel檔案: {filename}")
        print(f"  已設定自動換行")


def main():
    """主程式"""
    print("""
    ╔════════════════════════════════════════════════════════════╗
    ║       TASS 新聞爬蟲 (改進版 - 雙翻譯引擎)               ║
    ╚════════════════════════════════════════════════════════════╝
    """)
    
    urls = [
        'https://tass.ru/politika',
        'https://tass.ru/ekonomika',
        'https://tass.ru/armiya-i-opk',
        'https://tass.ru/mezhdunarodnaya-panorama',
    ]
    
    # 建立爬蟲（headless=False 可以看到瀏覽器運作）
    scraper = TASSNewsScraper(headless=True)
    
    # 設定每個類別抓取 20 篇文章
    max_per_category = 20
    
    print(f"\n將從 {len(urls)} 個類別各抓取最多 {max_per_category} 篇文章")
    print(f"預計總文章數: {len(urls) * max_per_category} 篇")
    print(f"預估時間: 爬取 40-60 分鐘 + 翻譯 60-100 分鐘 = 總計約 100-160 分鐘\n")
    
    try:
        # 爬取所有類別
        articles = scraper.scrape_all_categories(urls, max_per_category)
        
        if not articles:
            print("\n❌ 未爬取到任何文章！")
            return
        
        print(f"\n✓ 成功爬取 {len(articles)} 篇文章")
        
        # 自動進行翻譯（不詢問）
        print("\n" + "="*70)
        print("開始自動翻譯...")
        print("="*70)
        
        try:
            scraper.translate_articles(articles)
        except ImportError as e:
            print(f"\n⚠️ 缺少翻譯套件: {e}")
            print("請執行:")
            print("  pip install googletrans==4.0.0-rc1")
            print("  pip install deep-translator")
            # 翻譯失敗時填入空值
            for article in articles:
                article['title_zh'] = ''
                article['content_zh'] = ''
        except Exception as e:
            print(f"\n⚠️ 翻譯過程發生錯誤: {e}")
            # 確保所有文章都有翻譯欄位
            for article in articles:
                if 'title_zh' not in article:
                    article['title_zh'] = ''
                if 'content_zh' not in article:
                    article['content_zh'] = ''
        
        # 儲存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'tass_news_{timestamp}.xlsx'
        
        scraper.save_to_excel(articles, filename)
        
        print(f"""
    ╔════════════════════════════════════════════════════════════╗
    ║                     爬取完成！                            ║
    ╠════════════════════════════════════════════════════════════╣
    ║  文章數量: {len(articles):>3} 篇                                    ║
    ║  檔案名稱: {filename:<40} ║
    ╚════════════════════════════════════════════════════════════╝
        """)
        
        # 統計
        print("\n各類別統計:")
        from collections import Counter
        category_count = Counter(article.get('category', '未知') for article in articles)
        for category, count in category_count.items():
            print(f"  {category}: {count} 篇")
    
    finally:
        # 確保瀏覽器關閉
        del scraper


if __name__ == '__main__':
    main()